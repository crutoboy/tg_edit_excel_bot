import openpyxl as xl
from config import EXCEL_PATH


def xl_read_data():
    file = xl.load_workbook(EXCEL_PATH, read_only=True)
    worksheet = file.worksheets[0]
    rows = list(worksheet.rows)[2:]
    file.close()
    return rows

def xl_read_cell(row: int, column: int):
    file = xl.load_workbook(EXCEL_PATH, read_only=True)
    worksheet = file.worksheets[0]
    value = worksheet.cell(row, column)
    file.close()
    return value

def xl_read_modificators():
    file = xl.load_workbook(EXCEL_PATH, read_only=True)
    worksheet = file.worksheets[0]
    row = list(worksheet[1])
    file.close()
    return row

def xl_read_headers():
    file = xl.load_workbook(EXCEL_PATH, read_only=True)
    worksheet = file.worksheets[0]
    row = list(worksheet[2])
    file.close()
    return row

def xl_read_users_translate():
    file = xl.load_workbook(EXCEL_PATH, read_only=True)
    worksheet = file.worksheets[1]
    rows = list(worksheet.rows) # получение списка строк
    get_values = lambda x: x.value # задание именной безымянной для получения парамметра value у переменной
    rows_values = [tuple(map(get_values, row)) for row in rows] # проходимся по каждой строке, затем проходимся по каждой ячейки и записываем значения
    translate = dict(rows_values)
    file.close()
    return translate

def xl_write_cell(row: int, column: int, value):
    file = xl.load_workbook(EXCEL_PATH)
    worksheet = file.worksheets[0]
    worksheet.cell(row, column, value)
    file.save(EXCEL_PATH)
    file.close()

def xl_filter_column(rows: list, index_column: int, value) -> list:
    filtered_rows = [row for row in rows if row[index_column] == value]
    return list(filtered_rows)

def xl_filter_column_by_user(rows: list, user_id: int):
    modificators = xl_read_modificators()
    user_translation = xl_read_users_translate()

    user_name = user_translation.get(user_id)
    name_column_index = modificators.index("имя")

    filtered_rows = xl_filter_column(rows, name_column_index, user_name)

    return filtered_rows